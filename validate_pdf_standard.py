# -*- coding: utf-8 -*-
"""
validate_pdf_standard.py

Valida se um PDF contém os elementos obrigatórios do padrão definido.
Uso:
    python validate_pdf_standard.py caminho/para/arquivo.pdf

Dependências:
    pip install -r requirements.txt

Biblioteca recomendada: PyMuPDF (fitz)
"""
from __future__ import annotations
import sys
import os
import re
from typing import List
import argparse
import shutil
import json
import time
# optional fuzzy matching library to improve OCR tolerance
try:
    from rapidfuzz import fuzz
except Exception:
    fuzz = None

# OCR / rendering tunables
# Factor used when rendering PDF pages into images (1.0 = 72 DPI, 2.0 = 144 DPI)
RENDER_SCALE = 2.0
# Tesseract settings (can be tuned)
OCR_PSM = 6
OCR_OEM = 3

# Caminho padrão da planilha quando nenhum argumento for passado
DEFAULT_XLSX_PATH = r"C:\Users\victor.vasconcelos\Documents\Leitor de PDF\material de limpeza.xlsx"
LEARN_DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'learned_patterns.json')


def locate_tesseract(user_provided: str | None = None) -> str | None:
    """Tenta localizar um executável tesseract e retorna o caminho completo ou None.

    Ordem de busca:
    1. Caminho fornecido pelo usuário (user_provided).
    2. tesseract no PATH (shutil.which).
    3. Locais comuns de Scoop, conda e pasta 'tools' e a pasta onde você extraiu o zip.
    """
    candidates = []
    if user_provided:
        candidates.append(user_provided)

    # verificar PATH
    which_path = shutil.which("tesseract")
    if which_path:
        candidates.append(which_path)

    # use USERPROFILE env var on Windows, fallback to home
    user = os.environ.get('USERPROFILE') or os.path.expanduser('~')
    # caminhos prováveis (Scoop, tools, conda, AppData)
    probable = [
        os.path.join(user, "scoop", "shims", "tesseract.exe"),
        os.path.join(user, "scoop", "apps", "tesseract", "current", "tesseract.exe"),
        os.path.join(user, "tools", "tesseract", "tesseract.exe"),
        os.path.join(user, "tools", "tesseract-5.4.0.20240606", "tesseract.exe"),
        os.path.join(user, "AppData", "Local", "Programs", "Tesseract-OCR", "tesseract.exe"),
        os.path.join(user, "miniconda3", "Library", "bin", "tesseract.exe"),
        os.path.join(user, "anaconda3", "Library", "bin", "tesseract.exe"),
        os.path.join(user, "Documents", "Leitor de PDF", "tesseract-5.4.0.20240606", "tesseract.exe"),
    ]
    candidates.extend(probable)

    for p in candidates:
        try:
            if p and os.path.isfile(p):
                return os.path.abspath(p)
        except Exception:
            continue

    return None

try:
    import fitz  # PyMuPDF
except Exception as e:
    print("Erro ao importar PyMuPDF (fitz). Instale com: pip install PyMuPDF")
    raise

try:
    from PIL import Image
except Exception:
    Image = None

try:
    import pytesseract
except Exception:
    pytesseract = None


def normalize_space(s: str) -> str:
    """Remove quebras de linha e comprime espaços em branco para facilitar buscas."""
    return re.sub(r"\s+", " ", s).strip()


def normalize_for_search(s: str) -> str:
    """Lowercase, remove accents and compress spaces for robust searching."""
    if s is None:
        return ""
    import unicodedata
    s2 = unicodedata.normalize('NFKD', s)
    s2 = ''.join(c for c in s2 if not unicodedata.combining(c))
    s2 = s2.lower()
    s2 = re.sub(r"\s+", " ", s2).strip()
    return s2


def find_all_occurrences(text: str, substring: str) -> int:
    # busca simples (case-insensitive + accent-insensitive)
    return normalize_for_search(text).count(normalize_for_search(substring))


def find_digit_clusters(text: str, required_digits: int = 11, window_size: int = 24) -> List[str]:
    """Return list of digit clusters that contain at least required_digits digits within a window of window_size characters.

    This helps detect CPFs that may be split by spaces or punctuation in OCR output.
    """
    digits_positions = [(m.start(), m.group()) for m in re.finditer(r"\d", text)]
    clusters: List[str] = []
    if not digits_positions:
        return clusters

    # sliding window over digit positions
    n = len(digits_positions)
    i = 0
    while i < n:
        j = i
        # expand window while within window_size chars from start
        start_pos = digits_positions[i][0]
        digits = [digits_positions[i][1]]
        j = i + 1
        while j < n and (digits_positions[j][0] - start_pos) <= window_size:
            digits.append(digits_positions[j][1])
            j += 1
        if len(digits) >= required_digits:
            cluster = ''.join(digits)
            clusters.append(cluster)
            # skip past this window
            i = j
        else:
            i += 1
    return clusters


def contains_ci(text: str, substring: str) -> bool:
    # case-insensitive + accent-insensitive containment
    return normalize_for_search(substring) in normalize_for_search(text)


def fuzzy_contains(text: str, pattern: str, threshold: int = 80) -> bool:
    """Tolerant fuzzy check: returns True if pattern approximately appears in text.

    If rapidfuzz is not available, falls back to normalized substring check.
    """
    if not pattern:
        return False
    norm_text = normalize_for_search(text)
    norm_pattern = normalize_for_search(pattern)
    if fuzz is None:
        return norm_pattern in norm_text
    # Make matching more permissive by lowering the effective threshold slightly
    # This reduces false negatives on noisy OCR.
    effective = max(30, threshold - 10)
    score = fuzz.token_set_ratio(norm_pattern, norm_text)
    return score >= effective


def validate_pdf(path: str) -> List[str]:
    if not os.path.isfile(path):
        return [f"Arquivo não encontrado: {path}"]

    try:
        text = extract_text(path)
    except Exception as e:
        return [f"Erro ao extrair texto: {e}"]
    missing = []

    # Use fuzzy and token-based checks for robustness against OCR noise
    # 1. Título Principal (mais permissivo)
    title = "DECLARAÇÃO DE RECEBIMENTO DE MATERIAL DE LIMPEZA"
    if not fuzzy_contains(text, title, threshold=60):
        # also try shorter variant without 'DE LIMPEZA'
        if not fuzzy_contains(text, "DECLARAÇÃO DE RECEBIMENTO DE MATERIAL", threshold=60):
            missing.append(f"Título Principal: {title}")

    # 2. Instituição (várias formas aceitáveis)
    if not (fuzzy_contains(text, "CESMAC", threshold=65)
            or fuzzy_contains(text, "Cebraspe", threshold=70)
            or fuzzy_contains(text, "Centro Universitário", threshold=65)):
        missing.append("Identificação da Instituição: CESMAC ou Cebraspe (ou semelhante)")

    # 3. Identificação do Exame (aceitar variantes / ENEM)
    if not (fuzzy_contains(text, "PROCESSO SELETIVO", threshold=60)
            or fuzzy_contains(text, "Exame Nacional do Ensino Médio", threshold=60)
            or contains_ci(text, "ENEM")):
        missing.append("Identificação do Exame: PROCESSO SELETIVO / ENEM (ou semelhante)")

    # 4. Estrutura de Seções: verificar presença de tokens relevantes por seção
    # Seção I: procurar por 'SEÇÃO I' ou 'DADOS DO LOCAL' ou 'LOCAL CEDIDO'
    if not (fuzzy_contains(text, "SEÇÃO I", threshold=70) or fuzzy_contains(text, "DADOS DO LOCAL", threshold=65) or fuzzy_contains(text, "LOCAL CEDIDO", threshold=65)):
        missing.append("Seção I ausente (aprox): 'DADOS DO LOCAL CEDIDO' / 'SEÇÃO I'")
    # Seção II: 'SEÇÃO II' ou 'CEDENTE' ou 'DADOS DO CEDENTE'
    if not (fuzzy_contains(text, "SEÇÃO II", threshold=70) or fuzzy_contains(text, "CEDENTE", threshold=70) or fuzzy_contains(text, "DADOS DO CEDENTE", threshold=65)):
        missing.append("Seção II ausente (aprox): 'DADOS DO CEDENTE' / 'SEÇÃO II'")
    # Seção III: 'SEÇÃO III' ou 'CESSIONÁRIO' ou 'DADOS DO CESSIONÁRIO' or 'MATERIAL'
    if not (fuzzy_contains(text, "SEÇÃO III", threshold=70) or fuzzy_contains(text, "CESSIONÁRIO", threshold=70) or fuzzy_contains(text, "DADOS DO CESSIONÁRIO", threshold=65) or fuzzy_contains(text, "MATERIAL", threshold=65)):
        missing.append("Seção III ausente (aprox): 'DADOS DO CESSIONÁRIO' / 'SEÇÃO III'")

    # 5. Labels-chave (aceitar com/sem dois pontos e com ruído de OCR)
    labels = ["UF", "Município", "Nome do Local", "Representante do Local", "CPF", "Coordenador de Local"]
    for lab in labels:
        found = False
        if contains_ci(text, lab + ':') or contains_ci(text, lab):
            found = True
        else:
            # short labels: lower threshold to be more permissive for OCR noise
            # make 'CPF' and 'UF' more permissive since they're short and often noisy
            if lab.upper() in ("CPF", "UF"):
                th = 55
            else:
                th = 65 if len(lab) <= 4 else 80
            if fuzzy_contains(text, lab, threshold=th):
                found = True
        if not found:
            missing.append(f"Etiqueta ausente (aprox): '{lab}'")

        # CPF presence check: we only need to ensure the CPF field is filled (not validate the number)
        cpf_present = False
        try:
            # 1) If 'CPF' label appears, check following text for any non-whitespace content
            for m in re.finditer(r"(?i)\bcpf\b", text):
                start = m.end()
                context = text[start:start+120]
                # if there's any digit or letter in the context, consider it filled
                if re.search(r"[0-9A-Za-zÀ-ÿ]", context):
                    cpf_present = True
                    break
            # 2) fallback: any obvious CPF-like pattern anywhere
            if not cpf_present:
                if re.search(r"\d{3}[\.\-\s]?\d{3}[\.\-\s]?\d{3}[\-\s]?\d{2}", text):
                    cpf_present = True
            # 3) fallback: any cluster of ~11 digits (handles OCR-split cases)
            if not cpf_present:
                clusters = find_digit_clusters(text, required_digits=9, window_size=120)
                if clusters:
                    cpf_present = True
        except Exception:
            cpf_present = False

        if not cpf_present:
            missing.append("Etiqueta 'CPF' ausente (campo em branco)")

    # 6. Texto legal - usar prefixo fuzzy mais permissivo
    legal_prefix = "O CENTRO BRASILEIRO DE PESQUISA EM AVALIAÇÃO E SELEÇÃO"
    if not fuzzy_contains(text, legal_prefix, threshold=55):
        # try shorter tokens
        if not (fuzzy_contains(text, "CENTRO BRASILEIRO DE PESQUISA", threshold=60) or fuzzy_contains(text, "PESQUISA EM AVALIAÇÃO", threshold=60)):
            missing.append("Parágrafo legal obrigatório ausente (ou muito diferente)")

    # 7. Campos de assinatura (flexível)
    if not (fuzzy_contains(text, "CEDENTE", threshold=70) or fuzzy_contains(text, "CESSIONÁRIO", threshold=70) or contains_ci(text, "Assinatura") or contains_ci(text, "Assinado")):
        missing.append("Campo de assinatura ausente: 'CEDENTE'/'CESSIONÁRIO'/'Assinatura'")

    # If CPF specifically is missing, try a dedicated forced OCR pass and check regex directly
    try:
        if any('CPF' in m for m in missing):
            # quick regex on already-extracted text
            cpf_regex = re.compile(r"\d{3}[\.\-\s]?\d{3}[\.\-\s]?\d{3}[\-\s]?\d{2}")
            if not cpf_regex.search(text):
                extra_cpf_text = force_ocr_pdf_text(path, scale=max(2.5, RENDER_SCALE * 1.5), psm=11)
                if cpf_regex.search(extra_cpf_text):
                    missing = [m for m in missing if 'CPF' not in m]
    except Exception:
        pass

    # If critical items missing (title, CPF, sections), try a forced higher-resolution OCR pass
    critical_keywords = ('Título Principal', 'CPF', 'Seção ausente', 'Seção I ausente', 'Seção II ausente', 'Seção III ausente')
    if any(any(ck in m for ck in critical_keywords) for m in missing):
        try:
            extra = force_ocr_pdf_text(path, scale=max(2.5, RENDER_SCALE * 1.5), psm=11)
            if extra:
                # merge and re-check only the items that failed
                merged = text + "\n" + extra
                # re-check title
                if any('Título Principal' in m for m in missing):
                    if fuzzy_contains(merged, title, threshold=60) or fuzzy_contains(merged, "DECLARAÇÃO DE RECEBIMENTO DE MATERIAL", threshold=60):
                        missing = [m for m in missing if 'Título Principal' not in m]
                # re-check institution
                if any('Identificação da Instituição' in m for m in missing):
                    if (fuzzy_contains(merged, "CESMAC", threshold=65) or fuzzy_contains(merged, "Cebraspe", threshold=70)):
                        missing = [m for m in missing if 'Identificação da Instituição' not in m]
                # re-check exam
                if any('Identificação do Exame' in m for m in missing):
                    if (fuzzy_contains(merged, "PROCESSO SELETIVO", threshold=60) or fuzzy_contains(merged, "Exame Nacional do Ensino Médio", threshold=60) or 'ENEM' in merged.upper()):
                        missing = [m for m in missing if 'Identificação do Exame' not in m]
                # re-check sections
                new_sections = []
                for s in sections:
                    if fuzzy_contains(merged, s, threshold=65) or any(tok in merged for tok in ("SEÇÃO I", "DADOS DO LOCAL", "LOCAL CEDIDO")):
                        # if found, remove the corresponding missing entry
                        pass
                # re-check labels including CPF
                # CPF: recompute
                cpf_label_count = find_all_occurrences(merged, "CPF")
                cpf_numbers = re.findall(r"\d{3}[\.\-\s]?\d{3}[\.\-\s]?\d{3}[\-\s]?\d{2}", merged)
                cpf_clusters = find_digit_clusters(merged, required_digits=11, window_size=200)
                cfptotal = cpf_label_count + len(cpf_numbers) + len(cpf_clusters)
                if cfptotal >= 1:
                    missing = [m for m in missing if "CPF" not in m]
        except Exception:
            pass

    # Normalize and deduplicate missing items before returning.
    try:
        # Separate CPF-related messages and other messages
        cpf_related = [m for m in missing if 'cpf' in normalize_for_search(m)]
        other = [m for m in missing if 'cpf' not in normalize_for_search(m)]

        # Deduplicate other messages while preserving order
        seen = set()
        deduped_other = []
        for m in other:
            k = normalize_for_search(m)
            if k not in seen:
                seen.add(k)
                deduped_other.append(m)

        if cpf_related:
            # Choose a canonical CPF message. Prefer an existing detailed one if present.
            chosen = None
            for m in cpf_related:
                low = m.lower()
                if 'campo' in low or 'ausente' in low or 'etiqueta' in low:
                    chosen = m
                    break
            if not chosen:
                chosen = "Etiqueta 'CPF' ausente (campo em branco)"
            missing = deduped_other + [chosen]
        else:
            missing = deduped_other
    except Exception:
        # If anything goes wrong during normalization, fall back to the original list
        pass

    return missing


def ocr_image(pil_image: 'Image.Image') -> str:
    if pytesseract is None:
        raise RuntimeError("pytesseract não está instalado. Instale com: pip install pytesseract e instale o Tesseract-OCR no sistema.")
    try:
        # preprocess image and run tesseract with tuned config
        img = pil_image
        try:
            img = preprocess_for_ocr(img)
        except Exception:
            # fallback to original if preprocessing fails
            img = pil_image

        config = f"--psm {OCR_PSM} --oem {OCR_OEM}"
        # tentar com idioma português quando disponível
        try:
            return pytesseract.image_to_string(img, lang='por', config=config)
        except pytesseract.pytesseract.TesseractNotFoundError:
            raise RuntimeError("tesseract is not installed or it's not in your PATH. See README file for more information.")
        except Exception:
            # fallback sem idioma específico
            return pytesseract.image_to_string(img, config=config)
    except Exception:
        # repassar exceções para o chamador
        raise


def preprocess_for_ocr(pil_image: 'Image.Image') -> 'Image.Image':
    """Apply preprocessing steps to improve OCR accuracy:
    - convert to grayscale
    - upscale moderately
    - autocontrast
    - median filter to remove noise
    - Otsu thresholding (binarization)
    Returns a Pillow Image in 'L' or 'RGB' mode suitable for pytesseract.
    """
    try:
        img = pil_image.convert('L')
    except Exception:
        img = pil_image

    # upscale to help Tesseract on low-res images
    try:
        w, h = img.size
        scale = max(1.0, min(3.0, RENDER_SCALE))
        new_size = (int(w * scale), int(h * scale))
        if new_size != img.size:
            img = img.resize(new_size, resample=Image.BICUBIC)
    except Exception:
        pass

    # autocontrast to improve dynamic range; use lightweight filtering for speed
    try:
        from PIL import ImageOps, ImageFilter
        img = ImageOps.autocontrast(img)
        # median filter to reduce salt-and-pepper noise; smaller kernel to speed up
        try:
            img = img.filter(ImageFilter.MedianFilter(size=2))
        except Exception:
            pass
    except Exception:
        pass

    # Otsu thresholding (implemented via histogram)
    # Skip expensive histogram binarization for very large images to improve speed
    try:
        w, h = img.size
        if (w * h) <= 5000000:  # only run Otsu for images smaller than ~5MP
            hist = img.histogram()
            total = sum(hist)
            sumB = 0
            wB = 0
            maximum = 0.0
            sum1 = sum(i * hist[i] for i in range(256))
            level = None
            for i in range(256):
                wB += hist[i]
                if wB == 0:
                    continue
                wF = total - wB
                if wF == 0:
                    break
                sumB += i * hist[i]
                mB = sumB / wB
                mF = (sum1 - sumB) / wF
                between = wB * wF * (mB - mF) * (mB - mF)
                if between >= maximum:
                    level = i
                    maximum = between
            # apply threshold if computed
            if level is not None:
                img = img.point(lambda p: 255 if p > level else 0)
    except Exception:
        pass

    return img


def extract_text(path: str) -> str:
    """Extrai texto de um arquivo. Suporta PDFs e arquivos de imagem; para PDFs tenta extração nativa e, se insuficiente, usa OCR por página."""
    ext = os.path.splitext(path)[1].lower()
    image_exts = {'.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp'}

    if ext in image_exts:
        if Image is None:
            raise RuntimeError("Pillow não está instalado. Instale com: pip install pillow")
        img = Image.open(path)
        txt = ocr_image(img)
        return normalize_space(txt)

    # assumir PDF
    try:
        doc = fitz.open(path)
    except Exception as e:
        raise RuntimeError(f"Erro ao abrir PDF: {e}")

    pages_text: List[str] = []
    for p in doc:
        try:
            t = p.get_text("text")
        except Exception:
            t = ""

        # se a extração nativa for pequena, tentar OCR na página
        if t and len(t.strip()) > 30:
            pages_text.append(t)
        else:
            if Image is None or pytesseract is None:
                # sem OCR disponível, adicionar o texto nativo (mesmo vazia)
                pages_text.append(t or "")
            else:
                # renderizar página para imagem e aplicar OCR
                try:
                    mat = fitz.Matrix(RENDER_SCALE, RENDER_SCALE)
                    pix = p.get_pixmap(matrix=mat)
                    mode = "RGBA" if pix.alpha else "RGB"
                    img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
                    txt = ocr_image(img)
                    pages_text.append(txt)
                except Exception:
                    pages_text.append(t or "")

    raw_text = "\n".join(pages_text)
    return normalize_space(raw_text)


def force_ocr_pdf_text(path: str, scale: float = 3.0, psm: int | None = None) -> str:
    """Force OCR on all pages of a PDF using a specified render scale and optional psm.

    Returns the concatenated OCR text.
    """
    if Image is None or pytesseract is None:
        return ""
    try:
        doc = fitz.open(path)
    except Exception:
        return ""
    texts = []
    old_psm = None
    if psm is not None:
        try:
            old_psm = globals().get('OCR_PSM', None)
            globals()['OCR_PSM'] = psm
        except Exception:
            old_psm = None
    for p in doc:
        try:
            mat = fitz.Matrix(scale, scale)
            pix = p.get_pixmap(matrix=mat)
            mode = "RGBA" if pix.alpha else "RGB"
            img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
            txt = ocr_image(img)
            texts.append(txt)
        except Exception:
            continue
    if old_psm is not None:
        try:
            globals()['OCR_PSM'] = old_psm
        except Exception:
            pass
    return normalize_space("\n".join(texts))


def process_excel(xlsx_path: str, header_rows: int = 1, start_row: int | None = None, save_every: int = 10,
                  allowed_states: list | None = None) -> int:
    """Lê a planilha Excel e processa cada caminho listado na coluna G.

    Suporta reinício a partir de `start_row` e salva a cada `save_every` linhas para
    evitar perda de progresso em execuções longas com OCR.
    Retorna 0 se todos os PDFs estiverem aprovados; 1 caso contrário.
    """
    try:
        import openpyxl
    except Exception:
        print("Dependência ausente: openpyxl. Instale com: pip install openpyxl")
        return 2

    if not os.path.isfile(xlsx_path):
        print(f"Arquivo Excel não encontrado: {xlsx_path}")
        return 2

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    def load_learn_db() -> dict:
        try:
            if os.path.isfile(LEARN_DB):
                with open(LEARN_DB, 'r', encoding='utf-8') as fh:
                    return json.load(fh)
        except Exception:
            pass
        return {}

    def save_learn_db(d: dict):
        try:
            with open(LEARN_DB, 'w', encoding='utf-8') as fh:
                json.dump(d, fh, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def template_key_for_path(p: str) -> str:
        try:
            return os.path.basename(os.path.dirname(p)).strip().upper() or 'ROOT'
        except Exception:
            return 'ROOT'

    learn_db = load_learn_db()
    if learn_db:
        print(f"Aprendizado carregado: {len(learn_db)} entradas")

    # configure tesseract if available
    try:
        tcmd = locate_tesseract()
        if tcmd:
            try:
                import pytesseract as _pyt
                _pyt.pytesseract.tesseract_cmd = tcmd
                print(f"pytesseract configurado automaticamente para: {tcmd}")
            except Exception:
                pass
    except Exception:
        pass

    # Set TESSDATA_PREFIX if por.traineddata exists next to the script or in cwd
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        possible = [script_dir, os.getcwd(), os.path.dirname(xlsx_path)]
        found = None
        for d in possible:
            try:
                if os.path.isfile(os.path.join(d, 'por.traineddata')):
                    found = d
                    break
            except Exception:
                continue
        if found:
            os.environ['TESSDATA_PREFIX'] = found
            print(f"TESSDATA_PREFIX definido para: {found}")
    except Exception:
        pass

    results = []
    any_failed = False
    max_row = ws.max_row or 0

    # determine starting row
    if start_row is None:
        cur_row = header_rows + 1 if header_rows >= 0 else 1
    else:
        cur_row = max(int(start_row), header_rows + 1 if header_rows >= 0 else 1)

    if cur_row > max_row:
        print(f"Nenhuma linha para processar (start_row={cur_row}, max_row={max_row}).")
        return 2

    processed_since_save = 0

    # normalize allowed_states into a set for fast checks (uppercased)
    allowed_set = None
    try:
        if allowed_states:
            allowed_set = {str(s).strip().upper() for s in allowed_states}
            print(f"Filtrando somente UFs: {sorted(list(allowed_set))}")
    except Exception:
        allowed_set = None
    for r in range(cur_row, max_row + 1):
        # if allowed_set is provided, skip rows whose UF (column A) is not in the set
        if allowed_set is not None:
            try:
                uf_cell = ws.cell(row=r, column=1).value
                uf = (str(uf_cell).strip().upper() if uf_cell is not None else "")
                if uf not in allowed_set:
                    continue
            except Exception:
                continue

        cell_val = ws.cell(row=r, column=7).value
        if cell_val is None:
            continue

        pdf_path = str(cell_val).strip()
        if not pdf_path:
            continue

        pdf_path = os.path.expanduser(pdf_path)
        if not os.path.isabs(pdf_path):
            base_dir = os.path.dirname(os.path.abspath(xlsx_path))
            pdf_path = os.path.abspath(os.path.join(base_dir, pdf_path))

        print(f"\nProcessando linha {r}: {pdf_path}")

        tpl = template_key_for_path(pdf_path)
        preforced = False

        # auto-approve learned templates
        try:
            if tpl in learn_db and learn_db.get(tpl, {}).get('auto_approve'):
                # do not auto-approve when UF (col A) is missing
                uf_cell = ws.cell(row=r, column=1).value
                if uf_cell is None or str(uf_cell).strip() == '':
                    print(f"  -> Template '{tpl}' tem auto_approve mas UF ausente; não auto-aprovando (linha {r}).")
                else:
                    print(f"  -> Template aprendido '{tpl}' marcado como auto-approve; marcando como APROVADO.")
                    ws.cell(row=r, column=13).value = "Documento Aprovado"
                    try:
                        ws.cell(row=r, column=14).value = 'Auto-aprovado pelo aprendizado'
                    except Exception:
                        pass
                    results.append((pdf_path, True, []))
                    processed_since_save += 1
                    if save_every and processed_since_save >= int(save_every):
                        try:
                            wb.save(xlsx_path)
                            print(f"  -> Progresso salvo (linha {r}).")
                        except Exception as e:
                            print(f"  -> Falha ao salvar progresso: {e}")
                        processed_since_save = 0
                    continue
        except Exception:
            pass

        # apply learned forced OCR before validate if present
        if tpl in learn_db and learn_db.get(tpl, {}).get('always_force_ocr'):
            try:
                params = learn_db.get(tpl, {})
                scale = params.get('scale', max(2.5, RENDER_SCALE * 2.0))
                psm = params.get('psm', 6)
                print(f"  -> Aplicando OCR forçado antecipado (template aprendido: {tpl}) scale={scale} psm={psm}")
                extra = force_ocr_pdf_text(pdf_path, scale=scale, psm=psm)
                if extra:
                    orig_extract = globals().get('extract_text')
                    merged_text = (orig_extract(pdf_path) or '') + "\n" + extra
                    globals()['extract_text'] = lambda path: normalize_space(merged_text)
                    preforced = True
            except Exception as e:
                print('  -> Falha ao aplicar OCR forçado antecipado:', e)

        missing = validate_pdf(pdf_path)

        if preforced:
            try:
                if 'orig_extract' in locals() and orig_extract is not None:
                    globals()['extract_text'] = orig_extract
                else:
                    globals().pop('extract_text', None)
            except Exception:
                pass

        reprocess_note = None
        if missing and (Image is not None and pytesseract is not None):
            try:
                print("  -> Tentando OCR forçado (alta resolução) para reavaliar...")
                scale = max(2.5, RENDER_SCALE * 2.0)
                psm = 6
                extra = force_ocr_pdf_text(pdf_path, scale=scale, psm=psm)
                if extra:
                    try:
                        orig_extract = globals().get('extract_text')
                        merged_text = (orig_extract(pdf_path) or '') + "\n" + extra
                        globals()['extract_text'] = lambda path: normalize_space(merged_text)
                        missing_after = validate_pdf(pdf_path)
                    finally:
                        if orig_extract is not None:
                            globals()['extract_text'] = orig_extract
                        else:
                            globals().pop('extract_text', None)

                    if not missing_after:
                        print(" -> Documento APROVADO após OCR forçado. Padrão verificado.")
                        ws.cell(row=r, column=13).value = "Documento Aprovado"
                        try:
                            ws.cell(row=r, column=14).value = 'Aprovado após OCR forçado'
                        except Exception:
                            pass
                        results.append((pdf_path, True, []))

                        try:
                            key = template_key_for_path(pdf_path)
                            entry = learn_db.get(key, {})
                            entry.update({'always_force_ocr': True, 'psm': int(psm), 'scale': float(scale), 'last_seen': int(time.time())})
                            learn_db[key] = entry
                            save_learn_db(learn_db)
                            print(f"  -> Aprendizado salvo para template '{key}': always_force_ocr=True")
                        except Exception as e:
                            print('  -> Falha ao salvar aprendizado:', e)

                        processed_since_save += 1
                        if save_every and processed_since_save >= int(save_every):
                            try:
                                wb.save(xlsx_path)
                                print(f"  -> Progresso salvo (linha {r}).")
                            except Exception as e:
                                print(f"  -> Falha ao salvar progresso: {e}")
                            processed_since_save = 0
                        continue
                    else:
                        missing = missing_after
                        reprocess_note = 'Reprocessado: OCR forçado aplicado'
            except Exception as e:
                print('  -> Reprocessamento falhou:', e)

        if len(missing) == 0:
            print(" -> Documento APROVADO. Padrão verificado.")
            ws.cell(row=r, column=13).value = "Documento Aprovado"
            results.append((pdf_path, True, []))
        else:
            any_failed = True
            print(" -> Documento REPROVADO. Itens faltantes:")
            for item in missing:
                print("    - ", item)
            ws.cell(row=r, column=13).value = "Documento Reprovado"
            try:
                note = '; '.join(missing)
                if reprocess_note:
                    note = reprocess_note + ' - ' + note
                ws.cell(row=r, column=14).value = note
            except Exception:
                pass
            results.append((pdf_path, False, missing))

        processed_since_save += 1
        if save_every and processed_since_save >= int(save_every):
            try:
                wb.save(xlsx_path)
                print(f"  -> Progresso salvo (linha {r}).")
            except Exception as e:
                print(f"  -> Falha ao salvar progresso: {e}")
            processed_since_save = 0

    try:
        wb.save(xlsx_path)
        print(f"Planilha atualizada e salva: {xlsx_path}")
    except Exception as e:
        print(f"Erro ao salvar a planilha: {e}")
        return 2

    total = len(results)
    approved = sum(1 for r in results if r[1])
    failed = total - approved
    print("\nResumo:")
    print(f"  Processados: {total}")
    print(f"  Aprovados:   {approved}")
    print(f"  Reprovados:  {failed}")

    return 0 if (total > 0 and not any_failed) else (1 if total > 0 else 2)


def main(argv: List[str]):
    parser = argparse.ArgumentParser(
        description="Valida PDFs contra o padrão definido ou lê caminhos de uma planilha Excel (coluna G)."
    )
    parser.add_argument("path", nargs='?', default=None,
                        help="Caminho para um PDF ou para a planilha .xlsx. Se omitido, usa a planilha padrão do usuário.")
    parser.add_argument("--header-rows", type=int, default=1,
                        help="Número de linhas de cabeçalho a pular na planilha (padrão: 1)")
    parser.add_argument("--start-row", type=int, default=None,
                        help="Número da primeira linha a processar na planilha (1-based). Se omitido, inicia após header-rows.")
    parser.add_argument("--tesseract-cmd", type=str, default=None,
                        help="Caminho completo para o executável tesseract (ex: C:\\Users\\me\\tools\\tesseract\\tesseract.exe). Útil quando Tesseract não está no PATH.")
    args = parser.parse_args(argv[1:])

    path = args.path
    # Interatividade: se o usuário não fornecer caminho, perguntar explicitamente
    if path is None:
        try:
            prompt = input(f"Digite o caminho para a planilha (.xlsx) ou pressione Enter para usar o padrão [{DEFAULT_XLSX_PATH}]: ").strip()
        except Exception:
            prompt = ''
        if prompt:
            path = prompt
        else:
            # usar planilha padrão se existir
            print(f"Nenhum caminho especificado. Tentando planilha padrão: {DEFAULT_XLSX_PATH}")
            if os.path.isfile(DEFAULT_XLSX_PATH):
                path = DEFAULT_XLSX_PATH
            else:
                print(f"Arquivo padrão não encontrado: {DEFAULT_XLSX_PATH}")
                parser.print_usage()
                return 2

    # Configurar tesseract se fornecido via argumento
    if args.tesseract_cmd:
        try:
            import pytesseract as _pyt
            _pyt.pytesseract.tesseract_cmd = args.tesseract_cmd
            print(f"Configurado pytesseract para usar: {args.tesseract_cmd}")
        except Exception as e:
            print(f"Aviso: não foi possível configurar pytesseract com o caminho fornecido: {e}")
    else:
        # tentar localizar automaticamente; se não encontrado, pedir ao usuário o caminho
        auto = locate_tesseract()
        if auto:
            try:
                import pytesseract as _pyt
                _pyt.pytesseract.tesseract_cmd = auto
                print(f"Tesseract detectado automaticamente em: {auto}")
            except Exception:
                pass
        else:
            # interação: solicitar caminho do tesseract ao usuário
            print("Aviso: tesseract não foi encontrado automaticamente no sistema.")
            while True:
                try:
                    user_t = input("Digite o caminho completo para 'tesseract.exe' (ou pressione Enter para prosseguir sem Tesseract - não recomendado): ").strip()
                except Exception:
                    user_t = ''
                if not user_t:
                    print("Prosseguindo sem Tesseract. OCR não estará disponível e documentos escaneados poderão reprovar.")
                    break
                if os.path.isfile(user_t):
                    try:
                        import pytesseract as _pyt
                        _pyt.pytesseract.tesseract_cmd = user_t
                        print(f"Configurado pytesseract para usar: {user_t}")
                        break
                    except Exception as e:
                        print(f"Falha ao configurar pytesseract com o caminho informado: {e}")
                        # permitir nova tentativa
                        continue
                else:
                    print(f"Caminho inválido ou arquivo não encontrado: {user_t}")
                    # perguntar novamente
                    continue

    # Configurar TESSDATA_PREFIX se encontrarmos por.traineddata em locais prováveis
    # determinar caminho do executável que está sendo usado
    tesseract_cmd_used = None
    try:
        import pytesseract as _pyt2
        tesseract_cmd_used = getattr(_pyt2.pytesseract, 'tesseract_cmd', None)
    except Exception:
        tesseract_cmd_used = None

    # locais para procurar por.traineddata
    script_dir = os.path.dirname(os.path.abspath(__file__))
    cwd = os.getcwd()
    possible_tessdata_dirs = [
        os.path.join(script_dir),
        os.path.join(cwd),
        os.path.join(os.path.dirname(DEFAULT_XLSX_PATH)),
    ]
    # também checar tessdata ao lado do executável
    if tesseract_cmd_used and os.path.isfile(tesseract_cmd_used):
        exe_dir = os.path.dirname(tesseract_cmd_used)
        possible_tessdata_dirs.append(os.path.join(exe_dir, 'tessdata'))

    found = None
    for d in possible_tessdata_dirs:
        try:
            cand = os.path.join(d, 'por.traineddata')
            if os.path.isfile(cand):
                found = d
                break
        except Exception:
            continue

    if found:
        os.environ['TESSDATA_PREFIX'] = found
        print(f"TESSDATA_PREFIX definido para: {found} (por.traineddata detectado)")
    else:
        print("Aviso: não foi encontrado 'por.traineddata' em locais comuns.")
        print("Se o OCR falhar, coloque por.traineddata em a) <tesseract_install>/tessdata ou b) na pasta do script e reexecute.")

    if path.lower().endswith(('.xlsx', '.xls')):
        return process_excel(path, header_rows=args.header_rows, start_row=args.start_row)
    else:
        missing = validate_pdf(path)
        if len(missing) == 0:
            print("Documento APROVADO. Padrão verificado.")
            return 0
        else:
            print("Documento REPROVADO. Itens faltantes:")
            for item in missing:
                print(" - ", item)
                ws.cell(row=r, column=13).value = "Documento Reprovado"


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
